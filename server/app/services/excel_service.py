"""
Reads the "Available Accounts" pool from a local Excel file (.xlsx) instead
of Google Sheets -- no API, no credentials, no network call, no quota to
think about. Just point config.excel_accounts.file_path at a spreadsheet.

Expected layout (columns, in order, row 1 = header row and is skipped):
Account name | Phone No | Password | Remarks

To update the data: edit the .xlsx file directly (in Excel, LibreOffice, or
Google Sheets exported to .xlsx) and either wait for the cache to expire or
click Refresh on the dashboard, which bypasses the cache immediately.

Status badge inference: the file as specified has no dedicated "Status"
column, so status is inferred from keywords in the Remarks column. This is
inherently fragile (relies on consistent wording) -- if you want reliable
status tracking, add a 5th "Status" column with values like "Working" /
"In Use" / "Expired" and switch this module to read that column directly.
"""

from __future__ import annotations

import logging
import time
from typing import Any
import openpyxl
from app.config import get_config

logger = logging.getLogger("server.excel_service")

_cache: dict[str, Any] = {"data": None, "fetched_at": 0.0}

_EXPIRED_KEYWORDS = ("expired", "expire", "banned", "suspended", "dead", "invalid")
_IN_USE_KEYWORDS = ("in use", "in-use", "using", "assigned", "busy", "active session")


def _infer_status(remarks: str) -> str:
    lowered = (remarks or "").lower()
    for kw in _EXPIRED_KEYWORDS:
        if kw in lowered:
            return "expired"
    for kw in _IN_USE_KEYWORDS:
        if kw in lowered:
            return "in_use"
    return "working"


def _fetch_from_excel() -> list[dict]:
    cfg = get_config().excel_accounts



    workbook = openpyxl.load_workbook(cfg.file_path, read_only=True, data_only=True)
    sheet = workbook[cfg.sheet_name] if cfg.sheet_name in workbook.sheetnames else workbook.active

    accounts = []
    for row in sheet.iter_rows(min_row=cfg.header_rows + 1, max_col=4, values_only=True):
        account_name, phone_no, password, remarks = (row + (None, None, None, None))[:4]
        account_name = str(account_name).strip() if account_name is not None else ""
        if not account_name:
            continue  # skip fully blank rows
        phone_no = str(phone_no) if phone_no is not None else ""
        password = str(password) if password is not None else ""
        remarks = str(remarks) if remarks is not None else ""
        accounts.append(
            {
                "account_name": account_name,
                "phone_no": phone_no,
                "password": password,
                "remarks": remarks,
                "status": _infer_status(remarks),
            }
        )
    workbook.close()
    return accounts


def get_available_accounts(force_refresh: bool = False) -> dict:
    """
    Returns the cached (or freshly-read) account list. Raises whatever
    exception openpyxl raises on failure -- the router catches it and turns
    it into a clean {"success": false, "error": ...} response.
    """
    cfg = get_config().excel_accounts
    if not cfg.enabled:
        return {"success": False, "error": "Excel accounts source is not enabled in config.yaml"}

    now = time.time()
    if not force_refresh and _cache["data"] is not None and (now - _cache["fetched_at"]) < cfg.cache_seconds:
        accounts = _cache["data"]
    else:
        accounts = _fetch_from_excel()
        _cache["data"] = accounts
        _cache["fetched_at"] = now

    return {
        "success": True,
        "total_available": len(accounts),
        "accounts": accounts,
    }